import pygal 
import openpyxl

filename = 'data2.xlsx'

wb = openpyxl.load_workbook(filename)
sheets = wb.get_sheet_names()
sh_number = 1
for sh in sheets:
    sheet =  wb.get_sheet_by_name(sh)
    x_labels = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row - 2 + 1)]
    # x_labels = [r'$\mathregular{CO_2}$', r'$\mathregular{H_2S}$', 'CO', r'$\mathregular{O_2}$']
    y_labels = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row - 2 + 1)]
    t = sheet.cell(row=sheet.max_row - 1, column=2).value
    h = sheet.cell(row=sheet.max_row, column=2).value

    hist = pygal.HorizontalBar()
    hist.title = sheet.title
    hist.x_title = "t = {}°  h = {}%".format(t, h)
    hist.x_labels = x_labels
    hist.add("(мг/м3)", y_labels)
    hist.render_to_png('graphs2/{}.png'.format(sheet.title))
    sh_number +=1