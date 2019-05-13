import pygal 
import openpyxl

filename = 'data.xlsx'

wb = openpyxl.load_workbook(filename)
sheets = wb.get_sheet_names()

index = 1
for sh in sheets:
    sheet =  wb.get_sheet_by_name(sh)

    x_labels = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]
    y_labels = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row + 1)]

    y_labels = [int(i)/int(sheet.cell(row=1, column=3).value) for i in y_labels]
    print(y_labels)


    hist = pygal.Bar()
    hist.title = sheet.title
    hist.x_title = 'xi < x < xi+1'
    hist.x_labels = x_labels
    hist.add('n/h', y_labels)
    file_name = str(index) + '_graph'
    hist.render_to_png('math/{}.png'.format(file_name))
    index += 1
