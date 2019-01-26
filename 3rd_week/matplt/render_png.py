import pygal 
import openpyxl

wd = openpyxl.load_workbook('example.xlsx')
sheet =  wd.get_sheet_by_name('Балыкчы шаары')
x_labels = [sheet.cell(row=1, column=i).value for i in range(2, sheet.max_column + 1)]
time = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
col_1 = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)]
col_2 = [sheet.cell(row=i, column=3).value for i in range(2, sheet.max_row + 1)]
col_3 = [sheet.cell(row=i, column=4).value for i in range(2, sheet.max_row + 1)]
col_4 = [sheet.cell(row=i, column=4).value for i in range(2, sheet.max_row + 1)]
col_5 = [sheet.cell(row=i, column=6).value for i in range(2, sheet.max_row + 1)]

print(col_1)
print(col_2)
print(col_3)
print(time)
print(x_labels)


hist = pygal.Bar()
hist.title = sheet.title
hist.x_labels = time
hist.add(x_labels[0], col_1)
hist.add(x_labels[1], col_2)
hist.add(x_labels[2], col_3)
hist.add(x_labels[3], col_4)
hist.add(x_labels[4], col_5)
hist.render_to_png('graph.png')
