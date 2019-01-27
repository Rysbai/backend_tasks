import pygal 
import openpyxl

filename = 'data.xlsx'

wb = openpyxl.load_workbook(filename)
sheets = wb.get_sheet_names()

for sh in sheets:
    sheet =  wb.get_sheet_by_name(sh)
    x_labels = [sheet.cell(row=1, column=i).value for i in range(2, sheet.max_column + 1)]
    time = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]

    columns = []
    for col in range(2, sheet.max_column + 1):
        column = [sheet.cell(row=i, column=col).value for i in range(2, sheet.max_row + 1)]
        columns.append(column)

    hist = pygal.Bar()
    hist.title = sheet.title
    hist.x_title = "Өлчөө убактысы"
    hist.x_labels = time

    for i in range(len(columns)):
        hist.add(x_labels[i], columns[i])

    hist.render_to_png('graphs/{}.png'.format(sheet.title))
