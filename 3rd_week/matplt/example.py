import openpyxl
import matplotlib.pyplot as plt

wd = openpyxl.load_workbook('example.xlsx')
sheet =  wd.get_sheet_by_name('Лист1')
x_labels = [x for x in range(2, sheet.max_row)]
y_labels = [sheet.cell(row=i, column=2).value for i in x_labels]

plt.plot(x_labels, y_labels, c='yellow', linewidth=4)
plt.title("Балыкчы", fontsize=18)

x_labels = [x for x in range(2, sheet.max_row)]
y_labels = [sheet.cell(row=i, column=3).value for i in x_labels]
plt.plot(x_labels, y_labels, c='red', linewidth=4)

plt.show()